#Help
def scorecard():
	pass
from datetime import datetime
start_time = datetime.now()

###Code
####importing required libraries
import openpyxl
import pandas as pd
import os

#reading input files
india_inn = open("india_inns2.txt","r+") #india batting
pak_inn = open("pak_inns1.txt","r+") #pakistan batting
Playing_teams = open("teams.txt","r+")
input_team = Playing_teams.readlines()

pak_team = input_team[0]
pak_cricketers = pak_team[23:-1:].split(",")

ind_team = input_team[2]
ind_cricketers = ind_team[20:-1:].split(",")


lst_ind=india_inn.readlines() #124
for i in lst_ind:
    if i=='\n':
        lst_ind.remove(i)
      

lst_pak=pak_inn.readlines() #123
for i in lst_pak:
    if i=='\n':
        lst_pak.remove(i)

wb = openpyxl.Workbook()
sheet = wb.active

# batting [runs,ball,4s,6s,sr]
# bowling [over,medan,runs,Wickets, NB, WD, ECO]
#declaring required variables
Ind_out_count=0
FOW_pak=0
Pak_out_count={}
ind_bowlers={}
ind_bats={}

pak_batsman={}
pak_bowlers={}
pak_byes=0
Pak_bowlers_runs=0

########Pakistan Innings####################
for l in lst_pak:
    x=l.index(".")
    Pak_inn_overs=l[0:x+2]
    temp=l[x+2::].split(",")
    c_ball=temp[0].split("to") #0 2
    
    if f"{c_ball[0].strip()}" not in ind_bowlers.keys() :
        ind_bowlers[f"{c_ball[0].strip()}"]=[1,0,0,0,0,0,0]   #[over0,medan1,runs2,Wickets3, NB4, WD5, ECO6]
    elif "wide" in temp[1]:
        pass
    elif "byes" in temp[1]:                 #defining scores of byes
        if "FOUR" in temp[2]:
            pak_byes+=4
            ind_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "1 run" in temp[2]:
            pak_byes+=1
            ind_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "2 runs" in temp[2]:
            pak_byes+=2
            ind_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "3 runs" in temp[2]:
            pak_byes+=3
            ind_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "4 runs" in temp[2]:
            pak_byes+=4
            ind_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "5 runs" in temp[2]:
            pak_byes+=5
            ind_bowlers[f"{c_ball[0].strip()}"][0]+=1

    else:
        ind_bowlers[f"{c_ball[0].strip()}"][0]+=1
    
    if f"{c_ball[1].strip()}" not in pak_batsman.keys() and temp[1]!="wide":
        pak_batsman[f"{c_ball[1].strip()}"]=[0,1,0,0,0] #[runs,ball,4s,6s,sr]
    elif "wide" in temp[1] :
        pass
    else:
        pak_batsman[f"{c_ball[1].strip()}"][1]+=1
    

    if "out" in temp[1]:                           #updating scoresheet when out
        ind_bowlers[f"{c_ball[0].strip()}"][3]+=1
        if "Bowled" in temp[1].split("!!")[0]:
            Pak_out_count[f"{c_ball[1].strip()}"]=("b" + c_ball[0])
        elif "Caught" in temp[1].split("!!")[0]:
            w=(temp[1].split("!!")[0]).split("by")
            Pak_out_count[f"{c_ball[1].strip()}"]=("c" + w[1] +" b " + c_ball[0])
        elif "Lbw" in temp[1].split("!!")[0]:
            Pak_out_count[f"{c_ball[1].strip()}"]=("lbw  b "+c_ball[0])

    
       #updating scoresheet when run made by bat
    if "no run" in temp[1] or "out" in temp[1] :
        ind_bowlers[f"{c_ball[0].strip()}"][2]+=0
        pak_batsman[f"{c_ball[1].strip()}"][0]+=0
    elif "1 run" in temp[1]:
        ind_bowlers[f"{c_ball[0].strip()}"][2]+=1
        pak_batsman[f"{c_ball[1].strip()}"][0]+=1
    elif "2 runs" in temp[1]:
        ind_bowlers[f"{c_ball[0].strip()}"][2]+=2
        pak_batsman[f"{c_ball[1].strip()}"][0]+=2
    elif "3 runs" in temp[1]:
        ind_bowlers[f"{c_ball[0].strip()}"][2]+=3
        pak_batsman[f"{c_ball[1].strip()}"][0]+=3
    elif "4 runs" in temp[1]:
        ind_bowlers[f"{c_ball[0].strip()}"][2]+=4
        pak_batsman[f"{c_ball[1].strip()}"][0]+=4
    elif "FOUR" in temp[1]:
        ind_bowlers[f"{c_ball[0].strip()}"][2]+=4
        pak_batsman[f"{c_ball[1].strip()}"][0]+=4
        pak_batsman[f"{c_ball[1].strip()}"][2]+=1
    elif "SIX" in temp[1]:
        ind_bowlers[f"{c_ball[0].strip()}"][2]+=6
        pak_batsman[f"{c_ball[1].strip()}"][0]+=6
        pak_batsman[f"{c_ball[1].strip()}"][3]+=1
    elif "wide" in temp[1]:                   #updating scoresheet when wide 
        if "wides" in temp[1]:
            # print(temp[1][1])
            ind_bowlers[f"{c_ball[0].strip()}"][2]+=int(temp[1][1])
            ind_bowlers[f"{c_ball[0].strip()}"][5]+=int(temp[1][1])
        else:
            ind_bowlers[f"{c_ball[0].strip()}"][2]+=1
            ind_bowlers[f"{c_ball[0].strip()}"][5]+=1

for val in pak_batsman.values():
    val[-1]=round((val[0]/val[1])*100 , 2)


############# india innings ############## 
ind_bowlers_score=0
ind_byes=0

out_ind_bat={}
for l in lst_ind:
    x=l.index(".")
    over_ind=l[0:x+2]

    temp=l[x+2::].split(",")
 #updating scoresheet after byes####################
    c_ball=temp[0].split("to") #0 2
    if f"{c_ball[0].strip()}" not in pak_bowlers.keys() :
        pak_bowlers[f"{c_ball[0].strip()}"]=[1,0,0,0,0,0,0]   #[over0,medan1,runs2,Wickets3, NB4, WD5, ECO6]
    elif "wide" in temp[1]:
        pass
    elif "byes" in temp[1]:        
        if "FOUR" in temp[2]:
            ind_byes+=4
            pak_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "1" in temp[2]:
            ind_byes+=1
            pak_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "2" in temp[2]:
            ind_byes+=2
            pak_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "3" in temp[2]:
            ind_byes+=3
            pak_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "4" in temp[2]:
            ind_byes+=4
            pak_bowlers[f"{c_ball[0].strip()}"][0]+=1
        elif "5" in temp[2]:
            ind_byes+=5
            pak_bowlers[f"{c_ball[0].strip()}"][0]+=1
    else:
        pak_bowlers[f"{c_ball[0].strip()}"][0]+=1
    
    if f"{c_ball[1].strip()}" not in ind_bats.keys() and temp[1]!="wide":
        ind_bats[f"{c_ball[1].strip()}"]=[0,1,0,0,0] #[runs,ball,4s,6s,sr]
    elif "wide" in temp[1] :
        pass
    else:
        ind_bats[f"{c_ball[1].strip()}"][1]+=1
    
    #updating scoresheet after out#############
    if "out" in temp[1]:                    
        pak_bowlers[f"{c_ball[0].strip()}"][3]+=1
        
        if "Bowled" in temp[1].split("!!")[0]:
            out_ind_bat[f"{c_ball[1].strip()}"]=("b" + c_ball[0])
        elif "Caught" in temp[1].split("!!")[0]:
            w=(temp[1].split("!!")[0]).split("by")
            out_ind_bat[f"{c_ball[1].strip()}"]=("c" + w[1] +" b " + c_ball[0])
        elif "Lbw" in temp[1].split("!!")[0]:
            out_ind_bat[f"{c_ball[1].strip()}"]=("lbw  b "+c_ball[0])

    
#updating scoresheet after runs by bat###############    
    if "no run" in temp[1] or "out" in temp[1] :    
        pak_bowlers[f"{c_ball[0].strip()}"][2]+=0
        ind_bats[f"{c_ball[1].strip()}"][0]+=0
    elif "1 run" in temp[1]:
        pak_bowlers[f"{c_ball[0].strip()}"][2]+=1
        ind_bats[f"{c_ball[1].strip()}"][0]+=1
    elif "2 runs" in temp[1]:
        pak_bowlers[f"{c_ball[0].strip()}"][2]+=2
        ind_bats[f"{c_ball[1].strip()}"][0]+=2
    elif "3 runs" in temp[1]:
        pak_bowlers[f"{c_ball[0].strip()}"][2]+=3
        ind_bats[f"{c_ball[1].strip()}"][0]+=3
    elif "4 runs" in temp[1]:
        pak_bowlers[f"{c_ball[0].strip()}"][2]+=4
        ind_bats[f"{c_ball[1].strip()}"][0]+=4
    elif "FOUR" in temp[1]:
        pak_bowlers[f"{c_ball[0].strip()}"][2]+=4
        ind_bats[f"{c_ball[1].strip()}"][0]+=4
        ind_bats[f"{c_ball[1].strip()}"][2]+=1
    elif "SIX" in temp[1]:
        pak_bowlers[f"{c_ball[0].strip()}"][2]+=6
        ind_bats[f"{c_ball[1].strip()}"][0]+=6
        ind_bats[f"{c_ball[1].strip()}"][3]+=1
    elif "wide" in temp[1]:
        if "wides" in temp[1]:
            pak_bowlers[f"{c_ball[0].strip()}"][2]+=int(temp[1][1])
            pak_bowlers[f"{c_ball[0].strip()}"][5]+=int(temp[1][1])
        else:
            pak_bowlers[f"{c_ball[0].strip()}"][2]+=1
            pak_bowlers[f"{c_ball[0].strip()}"][5]+=1


for val in ind_bats.values():
    val[-1]=round((val[0]/val[1])*100 , 2)

for val in pak_batsman.values():
    val[-1]=round((val[0]/val[1])*100 , 2)

for val in ind_bowlers.values():
    if val[0]%6==0:
        val[0] = val[0]//6
    else:
        val[0] = (val[0]//6) + (val[0]%6)/10

for val in pak_bowlers.values():
    if val[0]%6==0:
        val[0] = val[0]//6
    else:
        val[0] = (val[0]//6) + (val[0]%6)/10

### Economy of ind bowlers
for val in ind_bowlers.values(): 
    x=str(val[0])
    if "." in x:
        balls = int(x[0])*6 + int(x[2])
        val[-1]=round((val[2]/balls)*6,1)
    else:
        val[-1] = round((val[2]/val[0]) ,1) 

### Economy of  pak bowlers
for val in pak_bowlers.values(): 
    x=str(val[0])
    if "." in x:
        balls = int(x[0])*6 + int(x[2])
        val[-1]=round((val[2]/balls)*6,1)
    else:
        val[-1] = round((val[2]/val[0]) ,1)

Fall_of_Wickets_paki="15-1(Babar Azam,2.4), 42-2(Fakhar Zaman,5.5), 87-3(Iftikhar Ahmed,12.1), 96-4(Rizwan,14.1), 97-5(Khushdil,14.3), 112-6(Asif Ali,16.3), 114-7(Mohammad Nawaz,17.1), 128-8(Shadab Khan,18.2), 128-9(Naseem Shah,18.3), 147-10(Dahani,19.5)"
Fall_of_Wickets_india = "1-1(Rahul,0.2), 50-2(Rohit,8.0), 53-3(Kohli,9.1), 89-4(Suryakumar Yadav,14.2), 136-5(Jadeja,19.1)"

# pakistan batting printing in scoresheet
Pak_batters_name=[]
for key in pak_batsman.keys():
    Pak_batters_name.append(key)


for i in range(len(pak_batsman)):
    sheet.cell(5+i,1).value = Pak_batters_name[i]
    sheet.cell(5+i,5).value = pak_batsman[Pak_batters_name[i]][0]
    sheet.cell(5+i,6).value = pak_batsman[Pak_batters_name[i]][1]
    sheet.cell(5+i,7).value = pak_batsman[Pak_batters_name[i]][2]
    sheet.cell(5+i,8).value = pak_batsman[Pak_batters_name[i]][3]
    sheet.cell(5+i,9).value = pak_batsman[Pak_batters_name[i]][4]
    if Pak_batters_name[i] not in Pak_out_count:
        sheet.cell(5+i,3).value = "not out"
    else:
        sheet.cell(5+i,3).value=Pak_out_count[Pak_batters_name[i]]


extra_data_paki = "5 (b 1, lb 0, w 4, nb 0, p 0)"
extra_data_india = "14 (b 0, lb 5, w 9, nb 0, p 0)"

sheet.cell(3,1).value = "Batter"
sheet["E3"] = "Runs"
sheet["F3"] = "Balls"
sheet["G3"] = " 4s "
sheet["H3"] = " 6s "
sheet["I3"] = "  SR  "

# Pakistan bowling printing in scoresheet
sheet["A21"] = "Bowler"
sheet["C21"] = "Over"
sheet["D21"] = "Maiden"
sheet["E21"] = "Runs"
sheet["F21"] = "Wicket"
sheet["G21"] = "No Ball"
sheet["H21"] = "Wide"
sheet["I21"] = "Economy"

Pak_bowlers_name=[]
for key in pak_bowlers.keys():
    Pak_bowlers_name.append(key)

for i in range(len(pak_bowlers)):
    sheet.cell(47+i,1).value = Pak_bowlers_name[i]
    sheet.cell(47+i,3).value = pak_bowlers[Pak_bowlers_name[i]][0]
    sheet.cell(47+i,4).value = pak_bowlers[Pak_bowlers_name[i]][1]
    sheet.cell(47+i,5).value = pak_bowlers[Pak_bowlers_name[i]][2]
    sheet.cell(47+i,6).value = pak_bowlers[Pak_bowlers_name[i]][3]
    sheet.cell(47+i,7).value = pak_bowlers[Pak_bowlers_name[i]][4]
    sheet.cell(47+i,8).value = pak_bowlers[Pak_bowlers_name[i]][5]
    sheet.cell(47+i,9).value = pak_bowlers[Pak_bowlers_name[i]][6]
    Pak_bowlers_runs+=pak_bowlers[Pak_bowlers_name[i]][2]
    Ind_out_count+=pak_bowlers[Pak_bowlers_name[i]][3]
