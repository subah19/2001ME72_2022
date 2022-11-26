from datetime import datetime
start_time = datetime.now()

#Help
def octant_analysis(mod=5000):
	pass


###Code
#importing required libraries
import os
import pandas as pd
import numpy as np
import math
import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills


# Mod rank count function
def octant_range_names(mod, filename):
    octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
    
    # Reading Excel File
    
    df = pd.read_excel(f'input\\{filename}')
    rows = df.shape[0]
    
    
    try:
        # Calculating Average Values
        u_avg = df['U'].mean()
        v_avg = df['V'].mean()
        w_avg = df['W'].mean()

        # Calculating Average Value of U, V, W
        df.insert(4, column="U Avg", value="")
        df.insert(5, column="V Avg", value="")
        df.insert(6, column="W Avg", value="")

        # Calculating U', V', W' 
        df.insert(7, column="U'=U - U avg", value="")
        df.insert(8, column="V'=V - V avg", value="")
        df.insert(9, column="W'=W - W avg", value="")

        df["U'=U - U avg"] = round(df['U'] - u_avg, 3)
        df["V'=V - V avg"] = round(df['V'] - v_avg, 3)
        df["W'=W - W avg"] = round(df['W'] - w_avg , 3)
        
        df.at[0, 'U Avg'] = round(u_avg, 3)
        df.at[0, 'V Avg'] = round(v_avg, 3)
        df.at[0, 'W Avg'] = round(w_avg, 3)
        
        df['U'] = round(df['U'], 3)
        df['V'] = round(df['V'], 3)
        df['W'] = round(df['W'], 3)
        df['T'] = round(df['T'], 3)
    except:
        print("Error in calculating average!")
        exit()
        
    try:
        #Inserting new column for Octant
        df.insert(10, column="Octant", value="")
        df.insert(11, column=" ", value="")
        df.insert(12, column="", value="")
        df.insert(13, column="Octant ID", value="")
        df.insert(14, column="1", value="")
        df.insert(15, column="-1", value="")
        df.insert(16, column="2", value="")
        df.insert(17, column="-2", value="")
        df.insert(18, column="3", value="")
        df.insert(19, column="-3", value="")
        df.insert(20, column="4", value="")
        df.insert(21, column="-4", value="")

        df.iloc[1, 12] = "Mod "+ str(mod)
        df.at[0, 'Octant ID'] = "Overall Count"
        l=[]
    except:
        print("Error in inserting columns.")
        exit()

    try:
        # Calculating the octant values
        for i in range(0, rows):
            if df.at[i,"U'=U - U avg"] >= 0 and  df.at[i,"V'=V - V avg"] >= 0:
                if df.at[i,"W'=W - W avg"] >= 0:
                  df.at[i, 'Octant'] = 1
                else:
                  df.at[i, 'Octant'] = -1
            elif df.at[i,"U'=U - U avg"] < 0 and  df.at[i,"V'=V - V avg"] >= 0:
                if df.at[i,"W'=W - W avg"] >= 0:
                  df.at[i, 'Octant'] = 2
                else:
                  df.at[i, 'Octant'] = -2
            elif df.at[i,"U'=U - U avg"] < 0 and  df.at[i,"V'=V - V avg"] < 0:
                if df.at[i,"W'=W - W avg"] >= 0:
                  df.at[i, 'Octant'] = 3
                else:
                  df.at[i, 'Octant'] = -3
            elif df.at[i,"U'=U - U avg"] >= 0 and  df.at[i,"V'=V - V avg"] < 0:
                if df.at[i,"W'=W - W avg"] >= 0:
                  df.at[i, 'Octant'] = 4
                else:
                  df.at[i, 'Octant'] = -4
            l.append(df.at[i, 'Octant'])

        df.at[0, "1"] = l.count(1)
        df.at[0, "-1"] = l.count(-1)
        df.at[0 ,"2"] = l.count(2)
        df.at[0 ,"-2"] = l.count(-2)
        df.at[0 ,"3"] = l.count(3)
        df.at[0 ,"-3"] = l.count(-3)
        df.at[0 ,"4"] = l.count(4)
        df.at[0 ,"-4"] = l.count(-4)
    except:
        print("Error in counting octant values.")
        exit()
    
    try:
        # Splitting list into ranges and finding the count of octant values
        start = 0
        end = len(l)
        step = int(mod)
        idx=1
        rows_tot_mod = math.ceil(rows/step)
        rows_tot = rows_tot_mod
        for i in range(start, end, step):
            x = i
            sub_list = l[x:x+step]
            y = x+step-1
            if y>rows:
                y=rows-1
            df.at[idx ,'Octant ID'] = str(x)+"-"+str(y)
            df.at[idx, '1'] = sub_list.count(1)
            df.at[idx, '-1'] = sub_list.count(-1)
            df.at[idx, '2'] = sub_list.count(2)
            df.at[idx, '-2'] = sub_list.count(-2)
            df.at[idx, '3'] = sub_list.count(3)
            df.at[idx, '-3'] = sub_list.count(-3)
            df.at[idx, '4'] = sub_list.count(4)
            df.at[idx, '-4'] = sub_list.count(-4)
            idx+=1
    except:
        print("Error in counting octant values for ranges!")
        exit()
    
    try:
        # Inserting Rank Columns 
        col_num = 22
        for i in range(1,5):
            header = "Rank Octant "+str(i)
            df.insert(col_num, column=header, value="")
            col_num+=1
            header = "Rank Octant "+str(-1*i)
            df.insert(col_num, column=header, value="")
            col_num+=1
        df.insert(col_num, column="Rank 1 Octant ID", value="")
        col_num+=1
        df.insert(col_num, column="Rank 1 Octant Name", value="")
        col_num+=1
        
        # Calculating rank for Overall Octant Count
        dict={}
        l=[]
        for i in range(1,5):
            oct_cnt = df.at[0, str(i)]
            dict[oct_cnt] = str(i)
            l.append(oct_cnt)
            oct_cnt = df.at[0, str(-1*i)]
            dict[oct_cnt] = str(-1*i)
            l.append(oct_cnt)
        
        l.sort(reverse=True)
        rank = 1
        df.at[0, "Rank 1 Octant ID"] = dict[l[0]]
        df.at[0, "Rank 1 Octant Name"] = octant_name_id_mapping[dict[l[0]]]
        
        for i in l:
            oct_id = "Rank Octant "+dict[i]
            df.at[0, oct_id] = rank
            rank+=1
        
        # Calculating rank for Mod Octant Count
        rank1=[]
        for idx in range(1, rows_tot_mod+1): 
            dict={}
            l=[]
            for i in range(1,5):
                oct_cnt = df.at[idx, str(i)]
                dict[oct_cnt] = str(i)
                l.append(oct_cnt)
                oct_cnt = df.at[idx, str(-1*i)]
                dict[oct_cnt] = str(-1*i)
                l.append(oct_cnt)

            l.sort(reverse=True)
            df.at[idx, "Rank 1 Octant ID"] = dict[l[0]]
            rank1.append(dict[l[0]])
            df.at[idx, "Rank 1 Octant Name"] = octant_name_id_mapping[dict[l[0]]]
            
            rank = 1
            for i in l:
                oct_id = "Rank Octant "+dict[i]
                df.at[idx, oct_id] = rank
                rank+=1
        
        # Count of Rank 1 Mod Values
        idx = rows_tot_mod+5
        df.at[idx, 'Rank Octant 4'] = "Octant ID"
        df.at[idx, 'Rank Octant -4'] = "Octant Name"
        df.at[idx, 'Rank 1 Octant ID'] = "Count of Rank 1 Mod Values"
        idx+=1
        for i in range(1,5):
            oct_id = str(i)
            cnt = rank1.count(oct_id)
            df.at[idx, 'Rank Octant 4'] = oct_id
            df.at[idx, 'Rank Octant -4'] = octant_name_id_mapping[oct_id]
            df.at[idx, 'Rank 1 Octant ID'] = cnt
            idx+=1
            
            oct_id = str(-1*i)
            cnt = rank1.count(oct_id)
            df.at[idx, 'Rank Octant 4'] = oct_id
            df.at[idx, 'Rank Octant -4'] = octant_name_id_mapping[oct_id]
            df.at[idx, 'Rank 1 Octant ID'] = cnt
            idx+=1
            
    except Exception as e:
        print("Error in calculating rank.", e)
            
    try:
        return df
    except:
        print("Error in exporting to CSV.")
        exit()
    
# Octant Transition Count function
def octant_transition_count(mod, df):
    try:
        # Reading Excel File
        rows = df.shape[0]
        step = mod
        cols = df.shape[1]
        df.insert(cols, column="                     ", value="")
        cols+=1
    except:
        print("Error in reading Excel file!")
        exit()
    
    try:
        # Overall Transition Count 
        for l in range(2,12):
            blank = ""
            for i in range(1,l+1):
                blank+=" "
            df.insert(cols, column=blank, value="")
            cols+=1
        blank_dict={}
        bl_len = 4        
        for i in range(1,5):
            blank=""
            for idx in range(0, bl_len):
                blank += " "
            blank_dict[str(i)] = blank
            bl_len+=1
            blank=""
            for idx in range(0, bl_len):
                blank += " "
            blank_dict[str(-1*i)] = blank
            bl_len+=1
        blank_dict['f'] = '  '
        blank_dict['s'] = '   '
        
        idx=0
        df.at[idx, blank_dict['1']] = 'To'
        idx+=1
        df.at[idx, blank_dict['s']] = 'Count'
        for k in range(-4,5):
            if k==0:
                continue
            df.at[idx, blank_dict[str(k)]] =  k
        idx+=1
        df.at[idx, blank_dict['f']] = "From"

        # Creating dataframe df1 to store values
        data=[]
        df1 = pd.DataFrame(data, index=['1','-1','2','-2','3','-3','4','-4'],
                        columns=['1','-1','2','-2','3','-3','4','-4'])

        df1 = df1.fillna(0)  # For filling 0 to df1

        # Calculating values
        for i in range(0,rows-1):
            first = str(df.at[i,'Octant'])
            second = str(df.at[i+1, 'Octant'])
            df1.at[first, second] += 1

        # Adding values to main dataframe
        for k in range (1,5):
            df.at[idx, blank_dict['s']] = str(k)
            for l in range (-4,5):
                if l==0:
                    continue
                df.at[idx, blank_dict[str(l)]] = df1.at[str(k), str(l)]
            idx+=1
            df.at[idx, blank_dict['s']] = str(-1*k)
            for l in range (-4,5):
                if l==0:
                    continue
                df.at[idx, blank_dict[str(l)]] = df1.at[str(-1*k), str(l)]
            idx+=1
    except Exception as e:
        print("Error in calculating Overall Transition Count!", e)
        exit()

    try:
        # Mod Transition Count
        for i in range(0, rows, step):
            lim = i+step
            if lim>=rows:
                lim = rows
            idx+=2
            df.at[idx, blank_dict['s']] = 'Mod Transition Count'
            idx+=1
            df.at[idx, blank_dict['s']] = str(i)+'-'+str(lim-1)
            df.at[idx, blank_dict['1']] = 'To'
            idx+=1
            df.at[idx, blank_dict['s']] = 'Octant #'
            for k in range(-4,5):
                if k==0:
                    continue
                df.at[idx, blank_dict[str(k)]] =  k
            idx+=1
            df.at[idx, blank_dict['f']] = "From"

            data=[]
            df1 = pd.DataFrame(data, index=['1','-1','2','-2','3','-3','4','-4'],
                            columns=['1','-1','2','-2','3','-3','4','-4'])
            df1 = df1.fillna(0)

            # Calculating values
            if lim==rows:
                lim-=1
            for j in range(i,lim):
                first = str(df.at[j,'Octant'])
                second = str(df.at[j+1, 'Octant'])
                df1.at[first, second] += 1

           # Adding values to the main dataframe
            for k in range (1,5):
                df.at[idx, blank_dict['s']] = str(k)
                for l in range (-4,5):
                    if l==0:
                        continue
                    df.at[idx, blank_dict[str(l)]] = df1.at[str(k), str(l)]
                idx+=1
                df.at[idx, blank_dict['s']] = str(-1*k)
                for l in range (-4,5):
                    if l==0:
                        continue
                    df.at[idx, blank_dict[str(l)]] = df1.at[str(-1*k), str(l)]
                idx+=1
    except:
        print("Error in calculating Mod Transition Count!")
        exit()
    
    try:
        # Exporting dataframe to excel
        return df
    except Exception as e:    
        print("Error in exporting to excel.", e)
        exit()

# Longest Subsequence function

def octant_longest_subsequence_count_with_range(mod, df, filename):
    try:
        # Reading Excel File
        rows = df.shape[0]
        cols = df.shape[1]
        df.insert(cols, column="                   ", value="")
        cols+=1
    except Exception as e:
        print("Error in reading Excel file!", e)
        exit()
    
    try:
        # Dataframe to store longest sequence and Count
        data=[]
        df1 = pd.DataFrame(data, index=['1','-1','2','-2','3','-3','4','-4'],
                       columns=['Len', 'Count'])
        df1 = df1.fillna(0)
        
        # Dataframe to store time ranges
        df3 = pd.DataFrame(data, columns=['1','-1','2','-2','3','-3','4','-4'])
        
        prev = df.at[0, 'Octant'] 
        df1.at[str(prev), 'Len'] = 1
        cur_len = 1
        ini = df.at[0,'T']
        fin = df.at[0, 'T']

        for idx in range(1,rows):
            cur = df.at[idx, 'Octant']
            if (cur == prev):
                cur_len+=1
            else:
                cur_len = 1
                ini = df.at[idx, 'T']
            fin = df.at[idx, 'T']
            df4 = df3.count(axis=0)
            if (cur_len == df1.at[str(cur), 'Len']):
                df1.at[str(cur), 'Count'] += 1                
                df3.at[df4[str(cur)], str(cur)] = ini
                df3.at[df4[str(cur)]+1, str(cur)] = fin
            elif(cur_len > df1.at[str(cur), 'Len']):
                df1.at[str(cur), 'Count'] = 1
                del df3[str(cur)]
                df3.insert(7, column = str(cur), value="")
                df3.at[0, str(cur)] = ini
                df3.at[1, str(cur)] = fin
            df3.replace('', np.nan, inplace=True)
            df1.at[str(cur), 'Len'] = max(cur_len, df1.at[str(cur), 'Len'])
            prev = cur

        # Inserting values in dataframe
        idx = 0
        for i in range(1,5):
            df.at[idx, 'Octant ##'] = str(i)
            df.at[idx, 'Longest Subsequence Length'] = df1.at[str(i), 'Len']
            df.at[idx, 'Count'] = df1.at[str(i), 'Count']
            idx+=1
            df.at[idx, 'Octant ##'] = str(-1*i)
            df.at[idx, 'Longest Subsequence Length'] = df1.at[str(-1*i), 'Len']
            df.at[idx, 'Count'] = df1.at[str(-1*i), 'Count']
            idx+=1
            
        cols=df.shape[1]
        df.insert(cols, column="                         ", value="")
        # Inserting df3 in main dataframe
        idx=0
        for i in range(1,5):
            
            # For positive i
            df.at[idx, 'Octant ###'] = str(i)
            df.at[idx, 'Longest Subsequence Length '] = df1.at[str(i), 'Len']
            df.at[idx, 'Count '] = df1.at[str(i), 'Count']
            idx+=1
            df.at[idx, 'Octant ###'] = "Time"
            df.at[idx, 'Longest Subsequence Length '] = "From"
            df.at[idx, 'Count '] = "To"
            idx+=1
            for index in range(0, len(df3[str(i)]), 2):
                if np.isnan(df3.at[index, str(i)]):
                    break
                df.at[idx, 'Longest Subsequence Length '] = df3.at[index, str(i)]
                df.at[idx, 'Count '] = df3.at[index+1, str(i)]
                idx+=1
            
            ### For negative i
            df.at[idx, 'Octant ###'] = str(-1*i)
            df.at[idx, 'Longest Subsequence Length '] = df1.at[str(-1*i), 'Len']
            df.at[idx, 'Count '] = df1.at[str(-1*i), 'Count']
            idx+=1
            df.at[idx, 'Octant ###'] = "Time"
            df.at[idx, 'Longest Subsequence Length '] = "From"
            df.at[idx, 'Count '] = "To"
            idx+=1
            
            for index in range(0, len(df3[str(-1*i)]), 2):
                if np.isnan(df3.at[index, str(-1*i)]):
                    break
                df.at[idx, 'Longest Subsequence Length '] = df3.at[index, str(-1*i)]
                df.at[idx, 'Count '] = df3.at[index+1, str(-1*i)]
                idx+=1     
    except Exception as e:
        print("Error in calculating longest sequence.", e)
        exit()
    
    try:
        # Exporting dataframe to excel 
        df.to_excel(f'output\\{filename[0:len(filename)-5]}vel_octant_analysis_mod{mod}.xlsx', index=False)
        return df
    except Exception as e:
        print("Error in exporting to Excel file!", e)
        exit()

# Octant analysis function
def octant_analysis(mod):
    dir_list = os.listdir('input') 
    for filename in dir_list:
        if(filename[-4:]=='xlsx'): 
            print(filename)
         
            df = octant_range_names(mod, filename)
            df = octant_transition_count(mod, df)
            df = octant_longest_subsequence_count_with_range(mod, df, filename)
                    
        path='input\\'+filename  
        outPath='output\\'+str(filename[0:len(filename)-5])+'vel_octant_analysis_mod'+str(mod)+'.xlsx'        
        worksheet = xl.load_workbook(path)
        sheet = worksheet.active
        fill_pattern = PatternFill(patternType="solid",fgColor="FFFF33")
        sheet['L1'].value=""
        sheet['AG1'].value=""
        sheet['AH1'].value=""
        sheet['AR1'].value=""
        tot_r = df.shape[0]
        tot_c = df.shape[1]
        rows_tot = math.ceil(tot_r/mod)
        r=0
        for row in sheet.iter_rows(min_row=2, min_col=1, max_row=tot_r+1, max_col=tot_c):
            c=0
            for cell in row: 
                cell.value = df.iat[r, c]
                c+=1
            r+=1
        for row in sheet.iter_rows(min_row=1, min_col=14, max_row=rows_tot+2, max_col=32):
            for cell in row: 
                # print(cell.value, end=" ")
                if(cell.value==1):
                    cell.fill = fill_pattern

        # Changing width of some columns
        sheet.column_dimensions['M'].width = 15
        sheet.column_dimensions['N'].width = 15
        sheet.column_dimensions['W'].width = 15
        sheet.column_dimensions['X'].width = 15
        sheet.column_dimensions['Y'].width = 15
        sheet.column_dimensions['Z'].width = 15
        sheet.column_dimensions['AA'].width = 15
        sheet.column_dimensions['AB'].width = 15
        sheet.column_dimensions['AC'].width = 15
        sheet.column_dimensions['AD'].width = 25
        sheet.column_dimensions['AE'].width = 24
        sheet.column_dimensions['AF'].width = 18
        sheet.column_dimensions['AT'].width = 24

        # Defining border formats 

        thin_border = Border(left=Side(border_style='thin',color='00000000'),
                        right=Side(border_style='thin',color='00000000'),
                        top=Side(border_style='thin',color='00000000'),
                        bottom=Side(border_style='thin',color='00000000')
                        )
        thick_border = Border(left=Side(border_style='thick',color='00000000'),
                    right=Side(border_style='thick',color='00000000'),
                    top=Side(border_style='thick',color='00000000'),
                    bottom=Side(border_style='thick',color='00000000')
                    )
                        
 
        col_num=19
        
        R_loc=1
        col_loc=14

        for i in range (R_loc,R_loc+rows_tot+2):
            for j in range (col_loc,col_loc+col_num):
                sheet.cell(row=i, column=j).border=thick_border

        for i in range (rows_tot+7,rows_tot+16):
            for j in range (29,32):
                sheet.cell(row=i, column=j).border=thick_border

        for row in sheet.iter_rows(min_row=4, min_col=36, max_row=11, max_col=43):
            for cell in row:
                cell.border = thick_border
        
        x=3
        for n in range(rows_tot+1):
            i=0
            for row in sheet.iter_rows(min_row=x, min_col=35, max_row=x+8, max_col=43):
                for cell in row:
                    if(cell.value!=None):
                        cell.border = thick_border
            x+=13

        for row in sheet.iter_rows(min_row=1, min_col=45, max_row=9, max_col=47):
            for cell in row:
                cell.border = thick_border

        max_rows = 1
        for ro in range(2, tot_r):
            
            if str(sheet.cell(row = ro, column = 50).value) == "nan":
                max_rows = ro
                break
        
        for row in sheet.iter_rows(min_row=1, min_col=49, max_row=max_rows-1, max_col=51):
            for cell in row:
                cell.border = thick_border
        
        sheet['E1']='U Avg'
        sheet['F1']='V Avg'
        sheet['G1']='W Avg'
        sheet['H1']="U'=U - U avg"
        sheet['I1']="V'=V - V avg"
        sheet['J1']="W'=W - W avg"
        sheet['K1']='Octant'
        sheet['N1']='Overall Octant Count'
        sheet['O1']='+1'
        sheet['P1']='-1'
        sheet['Q1']='+2'
        sheet['R1']='-2'
        sheet['S1']='+3'
        sheet['T1']='-3'
        sheet['U1']='+4'
        sheet['V1']='-4'
        sheet['W1']='Rank Octant 1'
        sheet['X1']='Rank Octant -1'
        sheet['Y1']='Rank Octant 2'
        sheet['Z1']='Rank Octant -2'
        sheet['AA1']='Rank Octant 3'
        sheet['AB1']='Rank Octant -3'
        sheet['AC1']='Rank Octant 4'
        sheet['AD1']='Rank Octant -4'
        sheet['AE1']='Rank1 Octant ID'
        sheet['AF1']='Rank1 Octant Name'
        sheet['AI1']='Overall Transition Count'
        sheet['AS1']='Longest Subsquence Length'
        sheet['AW1']='Longest Subsquence Length with Range'
        sheet['AX1']='Longest Subsquence Length'
        sheet['AY1']='Count'

        
        worksheet.save(outPath)

mod=5000
octant_analysis(mod)

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))